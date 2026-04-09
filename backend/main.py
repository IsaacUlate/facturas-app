from __future__ import annotations

import csv
import io
import json
import math
import re
import textwrap
import zipfile
import hashlib
from dataclasses import dataclass, asdict
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response, StreamingResponse
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from unidecode import unidecode


# =========================
# Config ARVOX por defecto
# =========================
BASE_DIR = Path(__file__).resolve().parent
ASSETS_DIR = BASE_DIR / "assets"
DOWNLOADED_INVOICES_FILE = BASE_DIR / "downloaded_invoices.json"

LOGO_CANDIDATES = [
    ASSETS_DIR / "arvox_logo.png",
    ASSETS_DIR / "arvox_logo.jpg",
    ASSETS_DIR / "arvox_logo.jpeg",
]

DEFAULT_ACCENT = "#8B2E00"
DEFAULT_EXCHANGE_RATE = 490.0
DEFAULT_SINPE = "8415-2881"
DEFAULT_FOOTER_NAME = "Arián Alfaro"


app = FastAPI(title="ARVOX Facturas API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


HEADER_ALIASES = {
    "customername": {"nombre", "cliente", "customer", "customername"},
    "miamicode": {"mia", "miami", "miacode", "miamicode"},
    "trackingnumber": {"guia", "guía", "tracking", "trackingnumber", "numero de guia", "numero guia"},
    "description": {
        "descripcion",
        "descripción",
        "producto",
        "productos",
        "item",
        "descripcion del producto",
        "product",
    },
    "status": {"estatus", "estado", "status"},
    "weightkg": {"peso kg", "peso", "pesokg", "kg"},
    "weightlb": {"peso libras", "peso lb", "peso libras.", "pesolb", "libras", "lb"},
    "priceperlb": {"precio por lb", "tarifa", "priceperlb", "precio lb", "tarifa lb", "precio"},
    "rowtotal": {"total", "monto", "rowtotal"},
}


# =========================
# Models
# =========================
@dataclass
class ParsedRow:
    row_number: int
    customer_name: str
    original_customer_name: str
    miami_code: str
    tracking_number: str
    description: str
    status: str
    weight_kg: Optional[float]
    weight_lb: Optional[float]
    price_per_lb: Optional[float]
    row_total: Optional[float]


@dataclass
class InvalidRow:
    row_number: int
    reason: str
    raw: Dict[str, Any]


# =========================
# Helpers
# =========================
def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_key(value: str) -> str:
    text = unidecode(normalize_text(value).lower())
    text = re.sub(r"[^a-z0-9 ]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def canonical_header(value: Any) -> Optional[str]:
    key = normalize_key(value)
    if not key:
        return None

    for canonical, aliases in HEADER_ALIASES.items():
        normalized_aliases = {normalize_key(alias) for alias in aliases}
        if key in normalized_aliases:
            return canonical

    if "peso" in key and ("libra" in key or key.endswith(" lb") or key == "lb"):
        return "weightlb"
    if "peso" in key and "kg" in key:
        return "weightkg"
    if "precio" in key and "lb" in key:
        return "priceperlb"

    return None


def parse_number(value: Any) -> Optional[float]:
    text = normalize_text(value)
    if not text:
        return None

    text = (
        text.replace("₡", "")
        .replace("$", "")
        .replace("USD", "")
        .replace("CRC", "")
        .replace("US$", "")
        .strip()
    )

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")

    try:
        return float(Decimal(text))
    except (InvalidOperation, ValueError):
        return None


def load_rows(file_bytes: bytes, filename: str) -> List[List[Any]]:
    lower = filename.lower()

    if lower.endswith(".csv"):
        decoded = file_bytes.decode("utf-8-sig", errors="replace")
        return [row for row in csv.reader(io.StringIO(decoded))]

    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        worksheet = workbook.worksheets[0]
        return [list(row) for row in worksheet.iter_rows(values_only=True)]

    if lower.endswith(".xls"):
        raise HTTPException(
            status_code=400,
            detail="Los archivos .xls antiguos no están soportados. Convierte el archivo a .xlsx o .csv.",
        )

    raise HTTPException(status_code=400, detail="Solo se permiten archivos CSV, XLSX o XLSM.")


def row_to_normalized_text(row: List[Any]) -> str:
    return normalize_key(" ".join(normalize_text(cell) for cell in row if normalize_text(cell)))


def is_completely_blank(row_values: List[Any]) -> bool:
    return all(not normalize_text(v) for v in row_values)


def normalize_customer_name(name: str) -> str:
    return normalize_key(name)


def find_pending_section_start(rows: List[List[Any]]) -> int:
    for idx, row in enumerate(rows[:120]):
        text = row_to_normalized_text(row)
        if "pendientes" in text:
            return idx
    raise HTTPException(status_code=400, detail="No encontré una sección 'PENDIENTES' en el archivo.")


def detect_header_row_from_index(rows: List[List[Any]], start_idx: int) -> Tuple[int, Dict[int, str]]:
    best_idx = -1
    best_map: Dict[int, str] = {}
    best_score = -1

    search_end = min(len(rows), start_idx + 20)

    for idx in range(start_idx, search_end):
        row = rows[idx]
        mapping: Dict[int, str] = {}
        score = 0

        for col_idx, value in enumerate(row):
            canonical = canonical_header(value)
            if canonical and canonical not in mapping.values():
                mapping[col_idx] = canonical
                score += 1

        has_customer = "customername" in mapping.values()
        has_tracking = "trackingnumber" in mapping.values()
        has_description = "description" in mapping.values()
        has_weight = "weightkg" in mapping.values() or "weightlb" in mapping.values()

        row_score = score
        if has_customer:
            row_score += 2
        if has_tracking:
            row_score += 2
        if has_description:
            row_score += 2
        if has_weight:
            row_score += 1

        if row_score > best_score and has_customer and has_tracking:
            best_idx = idx
            best_map = mapping
            best_score = row_score

    if best_idx == -1:
        raise HTTPException(
            status_code=400,
            detail="No pude detectar la fila de encabezados dentro de la sección PENDIENTES.",
        )

    return best_idx, best_map


def should_stop_section(normalized_full_row_text: str) -> bool:
    if not normalized_full_row_text:
        return False

    stop_keywords = [
        "factura",
        "facturas",
        "facturado",
        "facturados",
        "entregados",
        "entregado",
    ]

    if any(keyword in normalized_full_row_text for keyword in stop_keywords):
        return True

    if "pendientes" in normalized_full_row_text:
        return True

    return False


def build_rows(raw_rows: List[List[Any]]) -> Tuple[List[ParsedRow], List[InvalidRow]]:
    pending_start_idx = find_pending_section_start(raw_rows)
    header_idx, header_map = detect_header_row_from_index(raw_rows, pending_start_idx)

    parsed_rows: List[ParsedRow] = []
    invalid_rows: List[InvalidRow] = []

    blank_streak = 0

    for i in range(header_idx + 1, len(raw_rows)):
        row = raw_rows[i]
        normalized_full_row_text = row_to_normalized_text(row)

        if is_completely_blank(row):
            blank_streak += 1
            if blank_streak >= 5:
                break
            continue
        else:
            blank_streak = 0

        if should_stop_section(normalized_full_row_text):
            break

        extracted: Dict[str, Any] = {field: "" for field in HEADER_ALIASES.keys()}
        extracted.update({"weightkg": None, "weightlb": None, "priceperlb": None, "rowtotal": None})

        for col_idx, canonical in header_map.items():
            value = row[col_idx] if col_idx < len(row) else None
            if canonical in {"weightkg", "weightlb", "priceperlb", "rowtotal"}:
                extracted[canonical] = parse_number(value)
            else:
                extracted[canonical] = normalize_text(value)

        row_number = i + 1
        customer_name = extracted.get("customername", "")
        tracking_number = extracted.get("trackingnumber", "")
        description = extracted.get("description", "")

        raw_payload = {
            "row_number": row_number,
            "customer_name": customer_name,
            "tracking_number": tracking_number,
            "description": description,
        }

        if not customer_name and not tracking_number and not description:
            continue

        if not customer_name:
            invalid_rows.append(
                InvalidRow(
                    row_number=row_number,
                    reason="Falta NOMBRE",
                    raw=raw_payload,
                )
            )
            continue

        if not tracking_number:
            invalid_rows.append(
                InvalidRow(
                    row_number=row_number,
                    reason="Falta GUIA",
                    raw=raw_payload,
                )
            )
            continue

        parsed_rows.append(
            ParsedRow(
                row_number=row_number,
                customer_name=normalize_customer_name(customer_name),
                original_customer_name=customer_name,
                miami_code=extracted.get("miamicode", ""),
                tracking_number=tracking_number,
                description=description or "Sin descripción",
                status=extracted.get("status", ""),
                weight_kg=extracted.get("weightkg"),
                weight_lb=extracted.get("weightlb"),
                price_per_lb=extracted.get("priceperlb"),
                row_total=extracted.get("rowtotal"),
            )
        )

    return parsed_rows, invalid_rows


def get_effective_weight_lb(row: ParsedRow) -> Optional[float]:
    if row.weight_lb is not None:
        return round(float(row.weight_lb), 3)
    if row.weight_kg is not None:
        return round(float(row.weight_kg) * 2.20462, 3)
    return None


def get_price_per_lb_by_weight(weight_lb: Optional[float]) -> Optional[float]:
    if weight_lb is None:
        return None
    if float(weight_lb) < 1:
        return 6.99
    return 6.0


def calculate_row_total(row: ParsedRow, default_unit_price: float, default_price_per_lb: float) -> float:
    effective_weight_lb = get_effective_weight_lb(row)

    if effective_weight_lb is not None:
        if float(effective_weight_lb) < 1:
            return 6.99
        return round(float(effective_weight_lb) * 6.0, 2)

    if row.row_total is not None:
        return round(float(row.row_total), 2)

    return round(float(default_unit_price), 2)


def choose_best_customer_name(rows: List[ParsedRow]) -> str:
    return max(
        (r.original_customer_name.strip() for r in rows if r.original_customer_name.strip()),
        key=len,
        default="Cliente",
    )


def get_logo_path() -> Optional[Path]:
    for path in LOGO_CANDIDATES:
        if path.exists():
            return path
    return None


def format_tracking_last_6(tracking: str) -> str:
    tracking_text = normalize_text(tracking)
    digits = re.sub(r"\D", "", tracking_text)
    if len(digits) >= 6:
        return digits[-6:]
    if len(tracking_text) >= 6:
        return tracking_text[-6:]
    return tracking_text or "N/A"


def build_customer_invoices(
    rows: List[ParsedRow],
    default_unit_price: float,
    default_price_per_lb: float,
) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[ParsedRow]] = {}

    for row in rows:
        grouped.setdefault(row.customer_name, []).append(row)

    invoices: List[Dict[str, Any]] = []

    for customer_key, customer_rows in grouped.items():
        display_name = choose_best_customer_name(customer_rows)
        miami_codes = sorted({r.miami_code for r in customer_rows if r.miami_code})

        items: List[Dict[str, Any]] = []
        all_guides: List[str] = []

        for row in customer_rows:
            effective_weight_lb = get_effective_weight_lb(row)
            effective_price_per_lb = get_price_per_lb_by_weight(effective_weight_lb)

            if effective_weight_lb is not None:
                calculated_usd = calculate_row_total(row, default_unit_price, default_price_per_lb)
                if calculated_usd <= 0:
                    continue
                item_total_usd = round(float(calculated_usd), 2)
                item_total_crc = None
            elif row.row_total is not None and row.row_total > 0:
                item_total_crc = round(float(row.row_total), 2)
                item_total_usd = None
            else:
                calculated_usd = calculate_row_total(row, default_unit_price, default_price_per_lb)
                if calculated_usd <= 0:
                    continue
                item_total_usd = round(float(calculated_usd), 2)
                item_total_crc = None

            item_guides = [row.tracking_number] if row.tracking_number else []
            all_guides.extend(item_guides)

            items.append(
                {
                    "description": row.description,
                    "quantity": 1,
                    "weight_lb": effective_weight_lb,
                    "price_per_lb": effective_price_per_lb,
                    "total_usd": item_total_usd,
                    "total_crc": item_total_crc,
                    "guides": item_guides,
                }
            )

        if not items:
            continue

        unique_guides = sorted(set(all_guides))

        subtotal_crc = 0.0
        subtotal_usd = 0.0

        for item in items:
            subtotal_crc += float(item["total_crc"] or 0)
            subtotal_usd += float(item["total_usd"] or 0)

        invoices.append(
            {
                "customerKey": customer_key,
                "customerName": display_name,
                "guides": unique_guides,
                "miamiCodes": miami_codes,
                "items": items,
                "subtotal_crc": round(subtotal_crc, 2),
                "subtotal_usd": round(subtotal_usd, 2),
                "total_crc": round(subtotal_crc, 2),
                "total_usd": round(subtotal_usd, 2),
                "itemCount": len(items),
            }
        )

    invoices.sort(key=lambda x: x["customerName"].lower())
    return invoices


def money_usd(value: float) -> str:
    return f"${value:,.2f}"


def money_crc_text(value: float) -> str:
    return f"CRC {value:,.0f}"


# =========================
# Descargadas JSON
# =========================
def load_downloaded_state() -> Dict[str, Any]:
    if not DOWNLOADED_INVOICES_FILE.exists():
        return {"downloaded_invoice_ids": []}

    try:
        data = json.loads(DOWNLOADED_INVOICES_FILE.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            return {"downloaded_invoice_ids": []}
        if "downloaded_invoice_ids" not in data or not isinstance(data["downloaded_invoice_ids"], list):
            data["downloaded_invoice_ids"] = []
        return data
    except Exception:
        return {"downloaded_invoice_ids": []}


def save_downloaded_state(state: Dict[str, Any]) -> None:
    DOWNLOADED_INVOICES_FILE.write_text(
        json.dumps(state, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def build_invoice_identifier(invoice: Dict[str, Any]) -> str:
    normalized_items = []

    for item in invoice.get("items", []):
        normalized_items.append(
            {
                "description": normalize_key(item.get("description", "")),
                "guides": sorted(
                    format_tracking_last_6(g) for g in item.get("guides", []) if normalize_text(g)
                ),
                "weight_lb": round(float(item["weight_lb"]), 3) if item.get("weight_lb") is not None else None,
                "price_per_lb": round(float(item["price_per_lb"]), 2) if item.get("price_per_lb") is not None else None,
                "total_usd": round(float(item["total_usd"]), 2) if item.get("total_usd") is not None else None,
                "total_crc": round(float(item["total_crc"]), 2) if item.get("total_crc") is not None else None,
            }
        )

    normalized_items.sort(
        key=lambda x: (
            x["description"],
            ",".join(x["guides"]),
            x["weight_lb"] if x["weight_lb"] is not None else -1,
            x["total_usd"] if x["total_usd"] is not None else -1,
            x["total_crc"] if x["total_crc"] is not None else -1,
        )
    )

    payload = {
        "customer_name": normalize_customer_name(invoice.get("customerName", "")),
        "items": normalized_items,
    }

    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def filter_not_downloaded_invoices(invoices: List[Dict[str, Any]], downloaded_ids: set[str]) -> List[Dict[str, Any]]:
    filtered: List[Dict[str, Any]] = []

    for invoice in invoices:
        invoice_id = build_invoice_identifier(invoice)
        if invoice_id in downloaded_ids:
            continue
        filtered.append(invoice)

    return filtered


def mark_invoices_as_downloaded(invoices: List[Dict[str, Any]]) -> None:
    if not invoices:
        return

    state = load_downloaded_state()
    downloaded_ids = set(state.get("downloaded_invoice_ids", []))
    invoiced_guides = set(state.get("invoiced_guides", []))

    for invoice in invoices:
        invoice_id = build_invoice_identifier(invoice)
        downloaded_ids.add(invoice_id)
        for item in invoice.get("items", []):
            for guide in item.get("guides", []):
                g = normalize_text(guide)
                if g:
                    invoiced_guides.add(g)

    state["downloaded_invoice_ids"] = sorted(downloaded_ids)
    state["invoiced_guides"] = sorted(invoiced_guides)
    save_downloaded_state(state)


def create_invoice_pdf(invoice: Dict[str, Any], settings: Dict[str, Any]) -> bytes:
    buffer = io.BytesIO()
    page_width, page_height = A4
    c = canvas.Canvas(buffer, pagesize=A4)

    accent = colors.HexColor(settings.get("accentColor") or DEFAULT_ACCENT)
    exchange_rate = float(settings.get("exchangeRate") or DEFAULT_EXCHANGE_RATE)
    sinpe_number = settings.get("sinpeNumber") or DEFAULT_SINPE
    footer_name = settings.get("footerText") or DEFAULT_FOOTER_NAME

    customer_name = invoice.get("customerName", "Cliente")
    now = datetime.now(tz=__import__("zoneinfo").ZoneInfo("America/Costa_Rica"))
    invoice_date = f"{now.day}/{now.month}/{now.year}"

    items = invoice.get("items", [])
    if not items:
        items = [{
            "description": "Sin descripción",
            "weight_lb": None,
            "price_per_lb": None,
            "total_usd": None,
            "total_crc": 0.0,
            "guides": [],
        }]

    invoice_total_crc = float(invoice.get("total_crc") or 0)
    invoice_total_usd = float(invoice.get("total_usd") or 0)

    if invoice_total_usd <= 0 and invoice_total_crc > 0:
        invoice_total_usd = round(invoice_total_crc / exchange_rate, 2)
    if invoice_total_crc <= 0 and invoice_total_usd > 0:
        invoice_total_crc = round(invoice_total_usd * exchange_rate, 2)

    def centered(text: str, x: float, y: float, size=10, font="Helvetica", color=colors.black):
        c.setFont(font, size)
        c.setFillColor(color)
        c.drawCentredString(x, y, str(text))

    def left(text: str, x: float, y: float, size=10, font="Helvetica", color=colors.black):
        c.setFont(font, size)
        c.setFillColor(color)
        c.drawString(x, y, str(text))

    def right(text: str, x: float, y: float, size=10, font="Helvetica", color=colors.black):
        c.setFont(font, size)
        c.setFillColor(color)
        c.drawRightString(x, y, str(text))

    def pdf_money_usd(value: float) -> str:
        return f"${value:,.2f}"

    def pdf_money_crc_text(value: float) -> str:
        return f"CRC {value:,.2f}"

    def fit_text(text: str, max_len: int) -> str:
        t = str(text or "")
        return t if len(t) <= max_len else t[: max_len - 3] + "..."

    # ── Background ──
    c.setFillColor(colors.HexColor("#f5f5f5"))
    c.rect(0, 0, page_width, page_height, fill=1, stroke=0)

    # ── Accent border ──
    bm = 24
    c.setStrokeColor(accent)
    c.setLineWidth(9)
    c.rect(bm, bm, page_width - bm * 2, page_height - bm * 2, stroke=1, fill=0)

    # ── Content margins ──
    lx = 50   # left content edge
    rx = page_width - 50  # right content edge

    # ── Logo box — centered, larger ──
    logo_w = 220
    logo_h = 80
    logo_x = (page_width - logo_w) / 2
    logo_y = page_height - 52 - logo_h

    c.setFillColor(accent)
    c.roundRect(logo_x, logo_y, logo_w, logo_h, 12, stroke=0, fill=1)
    centered("ARVOX", page_width / 2, logo_y + 48, size=32, font="Helvetica-Bold", color=colors.white)
    centered("COURIER", page_width / 2, logo_y + 18, size=11, font="Helvetica", color=colors.white)

    # ── Client info row ──
    info_y = logo_y - 32
    c.setFont("Helvetica-Bold", 8)
    c.setFillColor(colors.HexColor("#888888"))
    c.drawString(lx, info_y, "CLIENTE")
    c.drawRightString(rx, info_y, "FECHA")

    c.setFont("Helvetica-Bold", 11)
    c.setFillColor(colors.black)
    c.drawString(lx, info_y - 16, fit_text(customer_name.upper(), 34))
    c.drawRightString(rx, info_y - 16, invoice_date)

    # ── Accent divider ──
    div1_y = info_y - 32
    c.setStrokeColor(accent)
    c.setLineWidth(1.5)
    c.line(lx, div1_y, rx, div1_y)

    # ── Table layout ──
    table_x = lx
    table_w = rx - lx

    col_guide = 88
    col_desc  = 196
    col_weight = 62
    col_price  = 74
    col_total  = table_w - col_guide - col_desc - col_weight - col_price

    cx_g = table_x + col_guide / 2
    cx_d = table_x + col_guide + col_desc / 2
    cx_w = table_x + col_guide + col_desc + col_weight / 2
    cx_p = table_x + col_guide + col_desc + col_weight + col_price / 2
    cx_t = table_x + col_guide + col_desc + col_weight + col_price + col_total / 2

    # ── Table header ──
    th_y = div1_y - 26
    th_h = 26

    c.setFillColor(accent)
    c.roundRect(table_x, th_y, table_w, th_h, 6, stroke=0, fill=1)

    centered("PAQUETE",      cx_g, th_y + 9, size=8, font="Helvetica-Bold", color=colors.white)
    centered("DESCRIPCIÓN",  cx_d, th_y + 9, size=8, font="Helvetica-Bold", color=colors.white)
    centered("PESO LB",      cx_w, th_y + 9, size=8, font="Helvetica-Bold", color=colors.white)
    centered("PRECIO / LB",  cx_p, th_y + 9, size=8, font="Helvetica-Bold", color=colors.white)
    centered("TOTAL",        cx_t, th_y + 9, size=8, font="Helvetica-Bold", color=colors.white)

    # ── Table rows ──
    row_y = th_y - 20
    row_gap = 22
    max_rows_visible = 9

    for idx, item in enumerate(items[:max_rows_visible]):
        if idx % 2 == 0:
            c.setFillColor(colors.HexColor("#ebebeb"))
            c.rect(table_x, row_y - 6, table_w, 20, fill=1, stroke=0)

        guide_text = (
            ", ".join(format_tracking_last_6(g) for g in item.get("guides", []))
            if item.get("guides") else "N/A"
        )
        description = fit_text((item.get("description") or "Sin descripción").upper(), 26)
        weight_lb   = item.get("weight_lb")
        price_per_lb = item.get("price_per_lb")

        if item.get("total_usd") is not None:
            item_total_usd = float(item["total_usd"])
        elif item.get("total_crc") is not None:
            item_total_usd = round(float(item["total_crc"]) / exchange_rate, 2)
        else:
            item_total_usd = 0.0

        centered(fit_text(guide_text, 10), cx_g, row_y, size=7, font="Helvetica")
        centered(description, cx_d, row_y, size=8, font="Helvetica")
        centered(
            "" if weight_lb is None else f"{float(weight_lb):.3f}".rstrip("0").rstrip("."),
            cx_w, row_y, size=8, font="Helvetica",
        )
        centered(
            "" if price_per_lb is None else pdf_money_usd(float(price_per_lb)),
            cx_p, row_y, size=8, font="Helvetica",
        )
        centered(pdf_money_usd(item_total_usd), cx_t, row_y, size=8, font="Helvetica")
        row_y -= row_gap

    # ── Divider after rows ──
    div2_y = row_y + 10
    c.setStrokeColor(colors.HexColor("#bbbbbb"))
    c.setLineWidth(0.8)
    c.line(table_x, div2_y, table_x + table_w, div2_y)

    # ── Total row ──
    tr_y = div2_y - 16
    centered("TOTAL", cx_g, tr_y, size=9, font="Helvetica-Bold")
    centered(pdf_money_usd(invoice_total_usd), cx_t, tr_y, size=9, font="Helvetica-Bold")

    # ── Summary section ──
    sum_div_y = tr_y - 26
    c.setStrokeColor(accent)
    c.setLineWidth(1.5)
    c.line(lx, sum_div_y, rx, sum_div_y)

    label_x = lx + 8
    val_x   = rx - 8

    usd_y = sum_div_y - 22
    left("TOTAL EN DÓLARES", label_x, usd_y, size=11, font="Helvetica-Bold", color=colors.HexColor("#444444"))
    right(pdf_money_usd(invoice_total_usd), val_x, usd_y, size=11, font="Helvetica-Bold")

    crc_y = usd_y - 22
    left("MONTO EN COLONES", label_x, crc_y, size=10, font="Helvetica", color=colors.HexColor("#666666"))
    right(pdf_money_crc_text(invoice_total_crc), val_x, crc_y, size=10, font="Helvetica")

    # ── "CANTIDAD A PAGAR" with highlighted box ──
    pay_label_y = crc_y - 36
    left("CANTIDAD A PAGAR", label_x, pay_label_y, size=13, font="Helvetica-Bold", color=colors.HexColor("#333333"))

    pay_box_w = 160
    pay_box_h = 34
    pay_box_x = rx - pay_box_w
    pay_box_y = pay_label_y - 10

    c.setFillColor(accent)
    c.roundRect(pay_box_x, pay_box_y, pay_box_w, pay_box_h, 8, stroke=0, fill=1)
    centered(
        pdf_money_crc_text(invoice_total_crc),
        pay_box_x + pay_box_w / 2,
        pay_box_y + 11,
        size=13, font="Helvetica-Bold", color=colors.white,
    )

    # ── Footer ──
    footer_mid_y = 72
    logo_path = get_logo_path()

    if logo_path:
        try:
            fl_w = 110
            fl_h = 110
            # Logo centered in the left half of the footer
            fl_x = page_width / 2 - fl_w - 18
            fl_y = footer_mid_y - fl_h / 2
            c.drawImage(
                ImageReader(str(logo_path)),
                fl_x, fl_y,
                width=fl_w, height=fl_h,
                preserveAspectRatio=True,
                mask="auto",
            )
            sinpe_cx = page_width / 2 + 60
        except Exception:
            sinpe_cx = page_width / 2
    else:
        sinpe_cx = page_width / 2

    centered("SINPE MÓVIL",  sinpe_cx, footer_mid_y + 26, size=11, font="Helvetica-Bold")
    centered(sinpe_number,   sinpe_cx, footer_mid_y + 4,  size=13, font="Helvetica-Bold", color=accent)
    centered(footer_name,    sinpe_cx, footer_mid_y - 18, size=10, font="Helvetica",      color=colors.HexColor("#666666"))

    c.showPage()
    c.save()

    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes
# =========================
# API
# =========================
@app.get("/api/health")
def health() -> Dict[str, str]:
    return {"status": "ok"}


@app.post("/api/process")
async def process_file(
    file: UploadFile = File(...),
    default_unit_price: float = Form(0.0),
    default_price_per_lb: float = Form(0.0),
):
    content = await file.read()
    raw_rows = load_rows(content, file.filename)
    rows, invalid_rows = build_rows(raw_rows)

    downloaded_state = load_downloaded_state()
    downloaded_ids = set(downloaded_state.get("downloaded_invoice_ids", []))
    invoiced_guides = set(downloaded_state.get("invoiced_guides", []))

    # Filter out individual rows whose guide was already invoiced
    rows = [r for r in rows if normalize_text(r.tracking_number) not in invoiced_guides]

    invoices = build_customer_invoices(rows, default_unit_price, default_price_per_lb)
    invoices = filter_not_downloaded_invoices(invoices, downloaded_ids)

    summary = {
        "totalRows": len(rows) + len(invalid_rows),
        "validRows": len(rows),
        "invalidRows": len(invalid_rows),
        "uniqueCustomers": len({r.customer_name for r in rows}),
        "invoicesToGenerate": len(invoices),
    }

    return {
        "summary": summary,
        "invoices": invoices,
        "invalidRows": [asdict(row) for row in invalid_rows],
    }


@app.post("/api/generate-pdf")
async def generate_pdf(payload: Dict[str, Any]):
    invoice = payload.get("invoice")
    settings = payload.get("settings", {})

    if not invoice:
        raise HTTPException(status_code=400, detail="Falta invoice en el payload.")

    downloaded_state = load_downloaded_state()
    downloaded_ids = set(downloaded_state.get("downloaded_invoice_ids", []))
    invoice_id = build_invoice_identifier(invoice)

    if invoice_id in downloaded_ids:
        raise HTTPException(status_code=400, detail="Esta factura ya fue descargada anteriormente.")

    pdf_bytes = create_invoice_pdf(invoice, settings)
    filename = normalize_key(invoice.get("customerName", "cliente")).replace(" ", "_") or "factura"
    headers = {"Content-Disposition": f'attachment; filename="factura_{filename}.pdf"'}

    mark_invoices_as_downloaded([invoice])

    return Response(content=pdf_bytes, media_type="application/pdf", headers=headers)


@app.post("/api/generate-zip")
async def generate_zip(payload: Dict[str, Any]):
    invoices = payload.get("invoices", [])
    settings = payload.get("settings", {})

    if not invoices:
        raise HTTPException(status_code=400, detail="No hay facturas para exportar.")

    downloaded_state = load_downloaded_state()
    downloaded_ids = set(downloaded_state.get("downloaded_invoice_ids", []))
    invoices_to_export = filter_not_downloaded_invoices(invoices, downloaded_ids)

    if not invoices_to_export:
        raise HTTPException(status_code=400, detail="No hay facturas nuevas para exportar. Todas ya fueron descargadas antes.")

    memory = io.BytesIO()

    with zipfile.ZipFile(memory, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for invoice in invoices_to_export:
            pdf_bytes = create_invoice_pdf(invoice, settings)
            name = normalize_key(invoice.get("customerName", "cliente")).replace(" ", "_") or "factura"
            zf.writestr(f"factura_{name}.pdf", pdf_bytes)

    mark_invoices_as_downloaded(invoices_to_export)

    memory.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="facturas.zip"'}

    return StreamingResponse(memory, media_type="application/zip", headers=headers)